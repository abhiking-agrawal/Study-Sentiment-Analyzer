import nltk

from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem.snowball import SnowballStemmer

stemmer = SnowballStemmer("english")

positiveWords = []
negativeWords = []
neutralWords = []
positiveWordsFileName = "./rt-polaritydata/pos.txt"
negativeWordsFileName = "./rt-polaritydata/neg.txt"
neutralWordsFileName = "./rt-polaritydata/neu.txt"

posF = open(positiveWordsFileName,"r")
negF = open(negativeWordsFileName,'r') 
# neuF = open(neutralWordsFileName,'r') 
for word in posF.readlines():
    positiveWords.append(word.rstrip('\n'))
for word in negF.readlines():
    negativeWords.append(word.rstrip('\n'))
# for word in neuF.readlines():
#     neutralWords.append(word.rstrip('\n'))

def getSentiment(obj):
    if obj["neg"] != 0:
        return "Negative"
        # if obj["neg"] >= obj["pos"]:
        #     return "Negative"
        # else:
        #     return "Positive"
    elif obj["pos"] != 0:
        return "Positive"
        # if obj["pos"] >= obj["neu"]:
        #     return "Positive"
        # else:
        #     return "Neutral"
    else:
        return "Neutral"

def computeSentiments(words):
    res = {"pos" : 0, "neg":0,"neu" : 0}
    for word in words:
        word = stemmer.stem(word)
        if word in positiveWords:
           # print(word + "==> pos")
            res["pos"] += 1
        elif word in negativeWords:
            #print(word + "==> neg")
            res["neg"] += 1
        else:
            #print(word + "==> neu")
            res["neu"] += 1

    return res
stop_words = set(stopwords.words('english'))

# print(computeSentiments(filtered_sentence))

#print(max(res, key=res.get))

# print(filtered_sentence)
exit = False
while exit:
    sentence = input("Enter a sentence.\n")
    sentence = sentence.lower()
    word_tokens = word_tokenize(sentence)
    filtered_sentence = [w for w in word_tokens if not w in stop_words]
    print(word_tokens)
    print(filtered_sentence)
    res = computeSentiments(filtered_sentence)
    print("==> " +getSentiment(res))

    flag = input("Do you wana exit? 0 - exit  1- continue\n")
    if flag == "0":
        exit = False

from textblob import TextBlob
from openpyxl import Workbook
from openpyxl import load_workbook
#wb = load_workbook(filename='rt-polaritydata/Manualvalidation_weather.xlsx')
wb = load_workbook(filename='rt-polaritydata/MeTooInput.xlsx')
ws = wb['Sheet1']


writeBook = Workbook()
dest_filename = 'output_book.xlsx'
rightSheet = writeBook.create_sheet(title="Sheet")
op = open("output.csv","w")

rowNo = 1
for row in ws.rows:
    #for cell in row:
    # print(row[5].value)
    month = row[3].value
    year = str(row[4].value)
    sentence = row[5].value.lower()
    tb = TextBlob(sentence)
    if tb.sentiment.polarity > 0:
        resultTextblob = "Positive"
    elif tb.sentiment.polarity == 0:
        resultTextblob = "Neutral"
    else:
        resultTextblob = "Negative"
    word_tokens = word_tokenize(sentence)
    filtered_sentence = [w for w in word_tokens if not w in stop_words]
    #print(word_tokens)
    #print(filtered_sentence)
    res = computeSentiments(filtered_sentence)
    result = getSentiment(res)
    # print(sentence + "==> " + result)
    # rightSheet.cell(column=1, row=rowNo, value= month)
    # rightSheet.cell(column=2, row=rowNo, value= year)
    # rightSheet.cell(column=3, row=rowNo, value= sentence)
    if rowNo == 1:
        op.write("Custom Analyzer"+","+"Textblob"+"\n")
        # rightSheet.cell(column=4, row=rowNo, value= "Custom Analyzer")
        # rightSheet.cell(column=5, row=rowNo, value= "Textblob")
    else:
        op.write(result+","+resultTextblob+"\n")
        # rightSheet.cell(column=4, row=rowNo, value= result)
        # rightSheet.cell(column=5, row=rowNo, value= resultTextblob)
    
    rowNo = rowNo + 1
    if rowNo % 10000 == 0:
        print(rowNo)

print(rowNo)
# writeBook.save(filename = dest_filename)
posF.close()
negF.close()
# neuF.close()
